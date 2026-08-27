"""Rapor CIKTILARI (P31) — Excel (xlsx) ve PDF uretimi.

UCU DE AYNI `RaporSonuc`TAN uretilir; tablo ciktisi zaten o nesnenin
kendisidir. Ayri ayri uretmek, ayni raporun uc yerde farkli rakam
gostermesine yol acardi.

KURUMSAL SABLON (PDF): site adi + aralik + zaman damgasi + sayfa altbilgisi
her sayfada. Logo OPSIYONELDIR ve YOKSA baslik metni kayar — logoyu zorunlu
kilmak, logo yuklememis bir siteye rapor urettirmemek olurdu.
"""
from __future__ import annotations

import io
from datetime import date, datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdf_canvas

from .raporlar import RaporSonuc, kurus_metin

#: PDF'te tek satira sigmayan sutun sayisi — ustunde YATAY sayfa kullanilir.
#: Dikey sayfada 9 sutun okunaksiz kucuk yaziya duser.
YATAY_ESIK = 6

#: (P181 Bölüm 8) Gömülü grafikte en fazla nokta. Web 60'a örnekliyor; gömülü
#: (statik, kağıt) grafikte daha az nokta okunur — üstünde eşit aralıkla
#: örneklenir (tarayıcı değil ama OKUNURLUK için).
MAX_GRAFIK_NOKTA = 30

#: Erişilebilir (renk körü dostu) palet — Okabe-Ito. Bilgi renk-YALNIZ değil:
#: çizgide her seriye AYRI İŞARET (marker), pastada etiket+yüzde, hepsinde
#: metin AÇIKLAMASI (legend) + eksen etiketi eşlik eder.
_GRAFIK_PALET = ("#0072B2", "#E69F00", "#009E73", "#CC79A7", "#56B4E9", "#D55E00")
#: Çizgi serisi işaretleri (reportlab makeMarker adları) — renkten bağımsız ayrım.
_PDF_ISARET = ("FilledCircle", "FilledSquare", "FilledDiamond",
               "FilledTriangleDown", "FilledStarFive", "FilledPentagon")
#: openpyxl çizgi işaret sembolleri.
_XLSX_ISARET = ("circle", "square", "diamond", "triangle", "star", "plus")


def _grafik_verisi(sonuc: RaporSonuc, grafik) -> tuple[list[str], list[tuple[str, list[float]]], bool] | None:
    """Rapor satırlarından grafik verisi: (etiketler, [(seri_adı, değerler)], örneklendi_mi).

    `grafik` ördek-tipli (`.tip/.x/.seriler` — katalogdaki `GrafikTanimi`;
    dairesel import olmasın diye tip import EDİLMEZ). Kuruş sütunları TL'ye
    çevrilir (okunur eksen). VERİ YOKSA (satır yok) `None` döner — çağıran
    boş grafik çizmez. Çok nokta varsa eşit aralıkla örneklenir.
    """
    if grafik is None or not sonuc.satirlar:
        return None
    tipler = {s.anahtar: s.tip for s in sonuc.sutunlar}
    basliklar = {s.anahtar: s.baslik for s in sonuc.sutunlar}

    def _sayi(anahtar: str, ham) -> float:
        if ham is None:
            return 0.0
        try:
            v = float(ham)
        except (TypeError, ValueError):
            return 0.0
        return v / 100 if tipler.get(anahtar) == "kurus" else v

    satirlar = list(sonuc.satirlar)
    n = len(satirlar)
    ornek = False
    if n > MAX_GRAFIK_NOKTA:
        adim = n / MAX_GRAFIK_NOKTA
        idx = sorted({min(n - 1, int(i * adim)) for i in range(MAX_GRAFIK_NOKTA)})
        satirlar = [satirlar[i] for i in idx]
        ornek = True

    etiketler = ["" if r.get(grafik.x) is None else str(r.get(grafik.x)) for r in satirlar]
    seriler: list[tuple[str, list[float]]] = []
    for a in grafik.seriler:
        seriler.append((basliklar.get(a, a), [_sayi(a, r.get(a)) for r in satirlar]))
    if not seriler:
        return None
    return etiketler, seriler, ornek


def _damga() -> str:
    return datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M UTC")


def _aralik_metni(baslangic: date | None, bitis: date | None) -> str:
    if baslangic and bitis:
        return f"{baslangic.strftime('%d.%m.%Y')} – {bitis.strftime('%d.%m.%Y')}"
    if bitis:
        return f"… – {bitis.strftime('%d.%m.%Y')}"
    if baslangic:
        return f"{baslangic.strftime('%d.%m.%Y')} – …"
    return "Tüm zamanlar"


def _excel_grafik(ws, sonuc: RaporSonuc, grafik, veri) -> None:
    """(P181 Bölüm 8) Yerel Excel grafiği göm (düzenlenebilir — cell referanslı).

    Grafik verisi TABLODAN AYRI, adanmış bir blokta (tablonun sağında) yazılır:
    örnekleme + kuruş→TL dönüşümü tabloda değil yalnız grafik için gerekli.
    Erişilebilirlik: legend (metin) + eksen başlığı; çizgide her seriye AYRI
    işaret; pastada kategori+yüzde etiketi. Başarısızlık raporu DÜŞÜRMEZ.
    """
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.marker import Marker
    from openpyxl.utils import get_column_letter

    etiketler, seriler, ornek = veri
    x_baslik = next((s.baslik for s in sonuc.sutunlar if s.anahtar == grafik.x), grafik.x)
    blok_sut = len(sonuc.sutunlar) + 2  # tablodan iki sütun boşluk sağda
    hdr = 5  # kaynak bloğu başlık satırı
    ws.cell(row=hdr, column=blok_sut, value=x_baslik + (" (örneklendi)" if ornek else ""))
    for j, (ad, _) in enumerate(seriler, start=1):
        ws.cell(row=hdr, column=blok_sut + j, value=ad)
    for i, et in enumerate(etiketler):
        ws.cell(row=hdr + 1 + i, column=blok_sut, value=et)
        for j, (_, vals) in enumerate(seriler, start=1):
            ws.cell(row=hdr + 1 + i, column=blok_sut + j, value=vals[i])
    son = hdr + len(etiketler)
    kats = Reference(ws, min_col=blok_sut, min_row=hdr + 1, max_row=son)

    if grafik.tip == "pasta":
        ch = PieChart()
        data = Reference(ws, min_col=blok_sut + 1, min_row=hdr, max_row=son)
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(kats)
        ch.dataLabels = DataLabelList()
        ch.dataLabels.showPercent = True
        ch.dataLabels.showCatName = True  # etiket renk-yalnız değil
    else:
        ch = LineChart() if grafik.tip == "cizgi" else BarChart()
        if grafik.tip == "sutun":
            ch.type = "col"
            ch.grouping = "clustered"
        data = Reference(ws, min_col=blok_sut + 1, max_col=blok_sut + len(seriler),
                         min_row=hdr, max_row=son)
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(kats)
        ch.x_axis.title = x_baslik
        ch.y_axis.title = "TL"
        ch.x_axis.delete = False
        ch.y_axis.delete = False
        if grafik.tip == "cizgi":
            for i, s in enumerate(ch.series):
                s.marker = Marker(symbol=_XLSX_ISARET[i % len(_XLSX_ISARET)], size=6)
                s.smooth = False  # işaret + düz çizgi: seriyi renksiz de ayırt et

    ch.title = sonuc.baslik
    ch.height = 8
    ch.width = 16
    ws.add_chart(ch, f"{get_column_letter(blok_sut)}{son + 3}")


# ================================= EXCEL ==================================== #
def excel_uret(
    sonuc: RaporSonuc, site_ad: str, baslangic: date | None, bitis: date | None,
    grafik=None,
) -> bytes:
    """XLSX uret — PARA HUCRELERI SAYIDIR, metin degil.

    Kurusu metin olarak yazmak, kullanicinin Excel'de toplam almasini
    engellerdi; rapor zaten Excel'e "uzerinde calisilsin" diye verilir.
    Hucre TL cinsinden `float` degil, KURUS/100 ondalikli sayidir ve
    bicimlendirme Excel'in kendi sayi bicimiyle yapilir.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sonuc.kod[:31] or "Rapor"

    ws.append([site_ad])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([sonuc.baslik])
    ws["A2"].font = Font(bold=True, size=12)
    ws.append([f"Dönem: {_aralik_metni(baslangic, bitis)}", f"Oluşturma: {_damga()}"])
    ws.append([])

    basliklar = [s.baslik for s in sonuc.sutunlar]
    ws.append(basliklar)
    for hucre in ws[ws.max_row]:
        hucre.font = Font(bold=True)
        hucre.alignment = Alignment(wrap_text=True, vertical="center")

    for satir in sonuc.satirlar:
        ws.append([_excel_deger(s, satir.get(s.anahtar)) for s in sonuc.sutunlar])

    if sonuc.toplamlar:
        ws.append([])
        toplam_satiri = []
        for i, s in enumerate(sonuc.sutunlar):
            if i == 0:
                toplam_satiri.append("TOPLAM")
            else:
                toplam_satiri.append(_excel_deger(s, sonuc.toplamlar.get(s.anahtar)))
        ws.append(toplam_satiri)
        for hucre in ws[ws.max_row]:
            hucre.font = Font(bold=True)

    # Para sutunlarina Turkce gruplama bicimi.
    for i, s in enumerate(sonuc.sutunlar, start=1):
        harf = get_column_letter(i)
        ws.column_dimensions[harf].width = max(12, min(len(s.baslik) + 4, 40))
        if s.tip == "kurus":
            for hucre in ws[harf]:
                if isinstance(hucre.value, (int, float)):
                    hucre.number_format = "#,##0.00"

    if sonuc.metin:
        ws.append([])
        ws.append([sonuc.metin])

    # (P181 Bölüm 8) Grafik göm (yapılandırılmışsa). Veri yoksa boş grafik
    # çizme — "veri yok" notu yaz. Grafik başarısızlığı raporu DÜŞÜRMEZ.
    if grafik is not None:
        veri = _grafik_verisi(sonuc, grafik)
        if veri is None:
            ws.append([])
            ws.append(["Grafik için veri yok."])
        else:
            try:
                _excel_grafik(ws, sonuc, grafik, veri)
            except Exception:
                pass

    tampon = io.BytesIO()
    wb.save(tampon)
    return tampon.getvalue()


def _excel_deger(sutun, ham):
    if ham is None:
        return None
    if sutun.tip == "kurus":
        # SAYI olarak yaz (TL cinsinden) — kullanici toplam alabilsin.
        return int(ham) / 100
    if sutun.tip == "tarih" and isinstance(ham, date):
        return ham
    return ham


# ================================== PDF ===================================== #
def pdf_uret(
    sonuc: RaporSonuc,
    site_ad: str,
    baslangic: date | None,
    bitis: date | None,
    logo_png: bytes | None = None,
    grafik=None,
) -> bytes:
    """Kurumsal sablonlu PDF.

    Sablon: site adi (+ logo varsa) · rapor adi · donem · zaman damgasi ·
    her sayfada "Sayfa n / m" altbilgisi. Sayfa sayisi ONCEDEN bilinmedigi
    icin iki gecisli uretim yapilir — "n / ?" yazmak resmi bir cikti icin
    kabul edilebilir degil.

    (P181 Bölüm 8) `grafik` verilir ve VERİ varsa son sayfada gömülü grafik
    çizilir (reportlab.graphics). Veri yoksa grafik sayfası eklenmez.
    """
    yatay = len(sonuc.sutunlar) > YATAY_ESIK
    sayfa = landscape(A4) if yatay else A4
    veri = _grafik_verisi(sonuc, grafik) if grafik is not None else None

    # 1. GECIS: sayfa sayisini ogren.
    toplam_sayfa = _ciz(sonuc, site_ad, baslangic, bitis, sayfa, logo_png, None, None, veri, grafik)
    # 2. GECIS: altbilgide gercek sayiyi yaz.
    tampon = io.BytesIO()
    _ciz(sonuc, site_ad, baslangic, bitis, sayfa, logo_png, tampon, toplam_sayfa, veri, grafik)
    return tampon.getvalue()


def _ciz(sonuc, site_ad, baslangic, bitis, sayfa, logo_png, tampon, toplam=None, veri=None, grafik=None):
    hedef = tampon or io.BytesIO()
    c = pdf_canvas.Canvas(hedef, pagesize=sayfa)
    genislik, yukseklik = sayfa
    kenar = 15 * mm
    satir_yuksekligi = 6 * mm

    sutunlar = sonuc.sutunlar
    birim = (genislik - 2 * kenar) / max(sum(s.genislik for s in sutunlar), 1)
    xler = []
    x = kenar
    for s in sutunlar:
        xler.append(x)
        x += s.genislik * birim

    sayfa_no = 0

    def _baslik() -> float:
        nonlocal sayfa_no
        sayfa_no += 1
        y = yukseklik - kenar
        if logo_png:
            try:
                from reportlab.lib.utils import ImageReader

                c.drawImage(
                    ImageReader(io.BytesIO(logo_png)), kenar, y - 14 * mm,
                    width=18 * mm, height=14 * mm, mask="auto",
                )
                sol = kenar + 22 * mm
            except Exception:
                # Bozuk/desteklenmeyen logo raporu DUSURMEZ: logo sussuz bir
                # ayrintidir, rapor ise gerekli.
                sol = kenar
        else:
            sol = kenar
        c.setFont("Helvetica-Bold", 14)
        c.drawString(sol, y - 5 * mm, site_ad)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(sol, y - 11 * mm, sonuc.baslik)
        c.setFont("Helvetica", 8)
        c.drawString(sol, y - 16 * mm, f"Dönem: {_aralik_metni(baslangic, bitis)}")
        c.drawRightString(genislik - kenar, y - 16 * mm, f"Oluşturma: {_damga()}")
        y -= 22 * mm
        c.setFont("Helvetica-Bold", 8)
        for s, sx in zip(sutunlar, xler):
            c.drawString(sx, y, s.baslik[:28])
        y -= 2 * mm
        c.line(kenar, y, genislik - kenar, y)
        return y - satir_yuksekligi

    def _altbilgi() -> None:
        c.setFont("Helvetica", 7)
        c.setFillColor(colors.grey)
        etiket = f"Sayfa {sayfa_no}" + (f" / {toplam}" if toplam else "")
        c.drawCentredString(genislik / 2, kenar / 2, etiket)
        c.setFillColor(colors.black)

    y = _baslik()
    c.setFont("Helvetica", 8)
    for satir in sonuc.satirlar:
        if y < kenar + 20 * mm:
            _altbilgi()
            c.showPage()
            y = _baslik()
            c.setFont("Helvetica", 8)
        for s, sx in zip(sutunlar, xler):
            ham = satir.get(s.anahtar)
            metin = (
                kurus_metin(ham) if s.tip == "kurus"
                else ("" if ham is None else str(ham))
            )
            if s.tip == "kurus":
                c.drawRightString(sx + s.genislik * birim - 3 * mm, y, metin)
            else:
                c.drawString(sx, y, metin[:40])
        y -= satir_yuksekligi

    if sonuc.toplamlar:
        y -= 2 * mm
        c.line(kenar, y + 4 * mm, genislik - kenar, y + 4 * mm)
        c.setFont("Helvetica-Bold", 8)
        for i, (s, sx) in enumerate(zip(sutunlar, xler)):
            if i == 0:
                c.drawString(sx, y, "TOPLAM")
            elif s.tip == "kurus":
                c.drawRightString(
                    sx + s.genislik * birim - 3 * mm, y,
                    kurus_metin(sonuc.toplamlar.get(s.anahtar)),
                )
        y -= satir_yuksekligi

    if sonuc.metin:
        c.setFont("Helvetica", 8)
        for parca in sonuc.metin.split("\n"):
            if y < kenar + 20 * mm:
                _altbilgi()
                c.showPage()
                y = _baslik()
                c.setFont("Helvetica", 8)
            c.drawString(kenar, y, parca[:160])
            y -= satir_yuksekligi

    # (P181 Bölüm 8) GRAFİK SAYFASI — veri varsa taze sayfada. Tablo satır
    # düzeniyle çakışmasın diye ayrı sayfa; başarısızlık raporu DÜŞÜRMEZ.
    if veri is not None:
        _altbilgi()
        c.showPage()
        yb = _baslik()
        try:
            _pdf_grafik(c, kenar, kenar + 12 * mm, genislik - 2 * kenar,
                        yb - (kenar + 14 * mm), sonuc, grafik, veri)
        except Exception:
            c.setFont("Helvetica", 9)
            c.drawString(kenar, yb, "Grafik çizilemedi.")

    _altbilgi()
    c.save()
    return sayfa_no


def _pdf_grafik(c, x, y, w, h, sonuc, grafik, veri) -> None:
    """(P181 Bölüm 8) reportlab.graphics ile gömülü grafik.

    Erişilebilirlik: renk TEK sinyal DEĞİL — çizgide her seriye AYRI işaret
    (marker), pastada dilim etiketi+yüzde, hepsinde legend (metin) + eksen
    başlığı. Erişilebilir (renk körü dostu) palet. `grafik.tip` çizim türü.
    """
    from reportlab.graphics import renderPDF
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.legends import Legend
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.widgets.markers import makeMarker

    etiketler, seriler, ornek = veri
    pal = [colors.HexColor(h) for h in _GRAFIK_PALET]
    d = Drawing(w, h)
    tip = grafik.tip
    x_baslik = next((s.baslik for s in sonuc.sutunlar if s.anahtar == grafik.x), grafik.x)

    # Başlık + örnekleme notu (metin — bilgi renk-yalnız değil).
    d.add(String(0, h - 12, sonuc.baslik, fontName="Helvetica-Bold", fontSize=11))
    if ornek:
        d.add(String(0, h - 26, f"(çok nokta — {len(etiketler)} noktaya örneklendi)",
                     fontName="Helvetica-Oblique", fontSize=8, fillColor=colors.grey))

    cizim_alt = 40  # legend/eksen için alt boşluk
    cizim_ust = h - 40
    cizim_yuk = cizim_ust - cizim_alt

    if tip == "pasta":
        ad0, vals0 = seriler[0]
        toplam = sum(vals0) or 1
        pie = Pie()
        pie.x, pie.y = w * 0.28, cizim_alt
        pie.width = pie.height = min(cizim_yuk, w * 0.44)
        pie.data = vals0
        pie.labels = [f"{et} %{round(v / toplam * 100)}" for et, v in zip(etiketler, vals0)]
        pie.sideLabels = True
        pie.slices.strokeWidth = 0.5
        for i in range(len(vals0)):
            pie.slices[i].fillColor = pal[i % len(pal)]
        d.add(pie)
        leg_pairs = [(pal[i % len(pal)], et[:22]) for i, et in enumerate(etiketler)]
        _pdf_legend(d, Legend, w * 0.62, cizim_ust, leg_pairs)
        renderPDF.draw(d, c, x, y)
        return

    if tip == "cizgi":
        ch = HorizontalLineChart()
    else:
        ch = VerticalBarChart()
    ch.x, ch.y = 44, cizim_alt
    ch.width, ch.height = w - 60, cizim_yuk
    ch.data = [vals for _, vals in seriler]
    ch.categoryAxis.categoryNames = etiketler
    ch.categoryAxis.labels.boxAnchor = "ne"
    ch.categoryAxis.labels.angle = 30
    ch.categoryAxis.labels.fontSize = 6
    ch.valueAxis.labels.fontSize = 6
    if tip == "cizgi":
        for i in range(len(seriler)):
            ch.lines[i].strokeColor = pal[i % len(pal)]
            ch.lines[i].strokeWidth = 1.5
            ch.lines[i].symbol = makeMarker(_PDF_ISARET[i % len(_PDF_ISARET)])
            ch.lines[i].symbol.size = 4
            ch.lines[i].symbol.fillColor = pal[i % len(pal)]
    else:
        for i in range(len(seriler)):
            ch.bars[i].fillColor = pal[i % len(pal)]
    d.add(ch)
    # Eksen başlıkları (metin).
    d.add(String(ch.x, 6, x_baslik, fontName="Helvetica", fontSize=8))
    d.add(String(0, cizim_ust + 4, "TL", fontName="Helvetica", fontSize=8))
    # Legend — seri adları (metin).
    leg_pairs = [(pal[i % len(pal)], ad[:22]) for i, (ad, _) in enumerate(seriler)]
    _pdf_legend(d, Legend, ch.x, cizim_ust + 16, leg_pairs)
    renderPDF.draw(d, c, x, y)


def _pdf_legend(d, Legend, x, y, pairs) -> None:
    leg = Legend()
    leg.x, leg.y = x, y
    leg.alignment = "right"
    leg.fontName = "Helvetica"
    leg.fontSize = 8
    leg.dxTextSpace = 4
    leg.deltay = 10
    leg.colorNamePairs = pairs
    d.add(leg)


def metin_pdf(baslik: str, govde: str, site_ad: str) -> bytes:
    """Serbest metin PDF (ihtar yazisi, denetim raporu).

    Tablo sablonundan AYRI: ihtar bir yazidir, sutunlu bir liste degil;
    tablo sablonuna sikistirmak metni hucrelere bolerdi.
    """
    tampon = io.BytesIO()
    c = pdf_canvas.Canvas(tampon, pagesize=A4)
    genislik, yukseklik = A4
    kenar = 20 * mm
    stil = getSampleStyleSheet()["BodyText"]
    y = yukseklik - kenar
    c.setFont("Helvetica-Bold", 12)
    c.drawString(kenar, y, baslik)
    y -= 10 * mm
    c.setFont("Helvetica", 10)
    for parca in govde.split("\n"):
        if y < kenar + 15 * mm:
            c.showPage()
            y = yukseklik - kenar
            c.setFont("Helvetica", 10)
        # Uzun satirlari kir: PDF kendiliginden sarmaz ve metin sayfadan
        # tasip GORUNMEZ olurdu.
        while len(parca) > 95:
            kes = parca.rfind(" ", 0, 95)
            kes = kes if kes > 40 else 95
            c.drawString(kenar, y, parca[:kes])
            parca = parca[kes:].lstrip()
            y -= 5 * mm
        c.drawString(kenar, y, parca)
        y -= 5 * mm
    c.setFont("Helvetica", 7)
    c.setFillColor(colors.grey)
    c.drawCentredString(genislik / 2, kenar / 2, f"{site_ad} · {_damga()}")
    c.save()
    _ = stil  # stil sayfasi ileride zengin metin icin; simdilik duz cizim.
    return tampon.getvalue()
